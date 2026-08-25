"""우체국 API로 받은 운송장번호를 어글리어스 송장 등록 페이지 업로드용 엑셀로 변환.

어글리어스 공급자 안내(2026-08-25)에 따르면, 주문번호/택배사/송장번호가 들어있는
엑셀을 그대로 업로드하면 시스템이 자동으로 인식해서 매칭한다. 위탁판매 채널 주문 중
실제 어글리어스 주문번호(숫자로만 된 상품주문번호 — uglyus_order_import가 채워줌)를
가진 것만 골라서 쓴다. 카카오톡 등으로 받은 다른 위탁판매 주문은 어글리어스 시스템과
무관하므로 제외된다.
"""
import os

import openpyxl

from schema import Channel

_SHEET_NAME = "송장등록"
_HEADERS = ["주문번호", "택배사", "송장번호"]
_COURIER_NAME = "우체국택배"


def write_uglyus_dispatch_excel(submit_results: list, output_path: str) -> int:
    """submit_results: epost_order_submit.submit_orders()가 반환하는 (package, response) 리스트.

    output_path가 이미 있으면 기존 행에 이어붙인다. 반환값은 이번 호출로 새로 쓴 행 수.
    """
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        ws = wb[_SHEET_NAME] if _SHEET_NAME in wb.sheetnames else wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _SHEET_NAME
        ws.append(_HEADERS)

    row_count = 0
    for package, response in submit_results:
        order = package.order
        if order.channel != Channel.WHOLESALE or not order.original_order_id.isdigit():
            continue
        regi_no = response.get("regiNo")
        if not regi_no or regi_no == "TESTREGINOAPI":
            continue
        ws.append([order.original_order_id, _COURIER_NAME, regi_no])
        row_count += 1

    wb.save(output_path)
    return row_count
