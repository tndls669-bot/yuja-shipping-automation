"""우체국 API로 받은 운송장번호를 네이버 스마트스토어 "엑셀 일괄발송" 양식으로 변환.

네이버 커머스API 연동 전까지 쓰는 다리 역할 — 스마트스토어 판매자센터의
[발주발송관리 > 엑셀 일괄발송]에 이 파일을 그대로 업로드하면 된다.
(안내문에 따르면 시트명이 반드시 "발송처리", 컬럼은 상품주문번호/배송방법/택배사/송장번호)
API 연동되면 이 모듈은 그대로 두고 daily_pipeline에 커머스API 호출 노드만 추가하면 된다.
"""
import os

import openpyxl

from schema import Channel

_SHEET_NAME = "발송처리"
_HEADERS = ["상품주문번호", "배송방법", "택배사", "송장번호"]
_DELIVERY_METHOD = "택배,등기,소포"
_COURIER_NAME = "우체국택배"


def write_naver_dispatch_excel(submit_results: list, output_path: str) -> int:
    """submit_results: epost_order_submit.submit_orders()가 반환하는 (package, response) 리스트.

    스마트스토어 채널 주문만 골라서 파일로 쓴다. output_path가 이미 있으면(하루에
    2단계를 여러 번 실행한 경우) 기존 행을 지우지 않고 이어붙인다. 반환값은 이번 호출로
    새로 쓴 행 수(누적 총 행 수가 아님).
    """
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        ws = wb[_SHEET_NAME] if _SHEET_NAME in wb.sheetnames else wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _SHEET_NAME
        ws.append(_HEADERS)

    row_count = 0
    for package, response in submit_results:
        order = package.order
        if order.channel != Channel.SMARTSTORE:
            continue
        regi_no = response.get("regiNo")
        if not regi_no or regi_no == "TESTREGINOAPI":
            continue
        ws.append([order.original_order_id, _DELIVERY_METHOD, _COURIER_NAME, regi_no])
        row_count += 1

    wb.save(output_path)
    return row_count

