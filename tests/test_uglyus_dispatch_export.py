import os
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import openpyxl  # noqa: E402

from courier_tier import Package  # noqa: E402
from schema import Channel, ProductGroup, build_standard_order  # noqa: E402
from uglyus_dispatch_export import write_uglyus_dispatch_excel  # noqa: E402


def _order(channel, order_id, name="이지은"):
    return build_standard_order(
        channel, order_id, ProductGroup.CHEONGYUJA, name, "01011112222", "주소", "11111", 1,
    )


def test_writes_wholesale_rows_with_numeric_order_id():
    results = [
        (Package(_order(Channel.WHOLESALE, "687372", "이지은"), "청유자", "2kg"), {"regiNo": "6890166281835"}),
    ]
    with tempfile.TemporaryDirectory() as tmp_dir:
        path = os.path.join(tmp_dir, "dispatch.xlsx")
        row_count = write_uglyus_dispatch_excel(results, path)

        assert row_count == 1
        wb = openpyxl.load_workbook(path)
        ws = wb["송장등록"]
        assert [c.value for c in ws[1]] == ["주문번호", "택배사", "송장번호"]
        assert [c.value for c in ws[2]] == ["687372", "우체국택배", "6890166281835"]


def test_wholesale_orders_without_real_numeric_id_are_excluded():
    # 카카오톡 등으로 받은 다른 위탁판매 주문은 어글리어스 자체 주문번호가 아니라
    # 우리 내부용 임시 ID(예: wholesale-20260825-001)이므로 제외해야 한다.
    results = [
        (Package(_order(Channel.WHOLESALE, "wholesale-20260825-001", "최규덕"), "청유자", "2kg"),
         {"regiNo": "6890166281829"}),
    ]
    with tempfile.TemporaryDirectory() as tmp_dir:
        path = os.path.join(tmp_dir, "dispatch.xlsx")
        row_count = write_uglyus_dispatch_excel(results, path)
        assert row_count == 0


def test_non_wholesale_orders_are_excluded():
    results = [
        (Package(_order(Channel.SMARTSTORE, "687372"), "청유자", "2kg"), {"regiNo": "6890166281835"}),
    ]
    with tempfile.TemporaryDirectory() as tmp_dir:
        path = os.path.join(tmp_dir, "dispatch.xlsx")
        row_count = write_uglyus_dispatch_excel(results, path)
        assert row_count == 0


def test_second_call_same_day_appends():
    with tempfile.TemporaryDirectory() as tmp_dir:
        path = os.path.join(tmp_dir, "dispatch.xlsx")
        write_uglyus_dispatch_excel(
            [(Package(_order(Channel.WHOLESALE, "111"), "청유자", "2kg"), {"regiNo": "a"})], path
        )
        write_uglyus_dispatch_excel(
            [(Package(_order(Channel.WHOLESALE, "222"), "청유자", "2kg"), {"regiNo": "b"})], path
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["송장등록"]
        assert ws.max_row == 3


if __name__ == "__main__":
    tests = [obj for name, obj in list(globals().items()) if name.startswith("test_")]
    for t in tests:
        t()
        print(f"PASS {t.__name__}")
    print(f"\n{len(tests)} tests passed")
