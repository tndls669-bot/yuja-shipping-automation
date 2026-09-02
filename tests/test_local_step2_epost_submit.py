import csv
import os
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import openpyxl  # noqa: E402

import local_step1_process_orders as step1  # noqa: E402
import local_step2_epost_submit as step2  # noqa: E402
from csv_export import write_orders_csv  # noqa: E402
from schema import Channel, ProductGroup, build_standard_order  # noqa: E402


def _set_epost_env():
    os.environ["EPOST_AUTH_KEY"] = "auth"
    os.environ["EPOST_SECURITY_KEY"] = "security"
    os.environ["EPOST_CUST_NO"] = "cust"
    os.environ["EPOST_APPR_NO"] = "appr"
    os.environ["EPOST_OFFICE_SER"] = "01"


def _write_pending(inbox_root, today, orders):
    day_dir = step1.today_dir(today, inbox_root)
    output_dir = os.path.join(day_dir, "output")
    os.makedirs(output_dir, exist_ok=True)
    write_orders_csv(orders, os.path.join(output_dir, f"{today}_orders.csv"), append=True)
    return output_dir


def test_no_pending_file_reports_nothing_to_submit():
    with tempfile.TemporaryDirectory() as inbox_root:
        original_today_dir = step2.today_dir
        step2.today_dir = lambda today: step1.today_dir(today, inbox_root)
        try:
            result = step2.run(today="2026-08-14", send_email=False)
        finally:
            step2.today_dir = original_today_dir
        assert "접수 대기 중인 주문이 없습니다" in result


def test_successful_submission_writes_result_csv_and_excludes_from_resubmission():
    order = build_standard_order(
        Channel.PHONE_TEXT, "phone-1", ProductGroup.YUJACHEONG,
        "홍길동", "01012345678", "전남 고흥군 ...", "59554", 2,
    )

    with tempfile.TemporaryDirectory() as inbox_root:
        output_dir = _write_pending(inbox_root, "2026-08-14", [order])
        _set_epost_env()

        original_today_dir = step2.today_dir
        original_insert_order = step2.insert_order
        step2.today_dir = lambda today: step1.today_dir(today, inbox_root)
        step2.insert_order = lambda auth_key, security_key, params: {"regiNo": "6890161633630"}
        try:
            result = step2.run(today="2026-08-14", send_email=False)
            # 오늘 주문 CSV는 리포트이자 접수 기록이라 지워지지 않고 그대로 남아있어야 함.
            assert os.path.exists(os.path.join(output_dir, "2026-08-14_orders.csv"))
            # 하지만 이미 접수완료로 기록됐으니, 다시 돌려도 재접수 대상에서 빠져야 함.
            second_result = step2.run(today="2026-08-14", send_email=False)
        finally:
            step2.today_dir = original_today_dir
            step2.insert_order = original_insert_order

        assert "6890161633630" in result
        assert "접수 대기 중인 주문이 없습니다" in second_result

        with open(os.path.join(output_dir, "2026-08-14_epost_result.csv"), encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        assert rows[1][0] == "phone-1"
        assert rows[1][1] == "홍길동"
        assert rows[1][4] == "6890161633630"
        assert rows[1][5] == "접수완료"


def test_api_failure_is_recorded_but_does_not_crash():
    order = build_standard_order(
        Channel.WHOLESALE, "wholesale-1", ProductGroup.CHEONGYUJA,
        "김철수", "01033334444", "전남 고흥군 ...", "59554", 7,
    )

    with tempfile.TemporaryDirectory() as inbox_root:
        output_dir = _write_pending(inbox_root, "2026-08-14", [order])
        _set_epost_env()

        def _fail(auth_key, security_key, params):
            raise RuntimeError("우체국 서버 오류")

        original_today_dir = step2.today_dir
        original_insert_order = step2.insert_order
        step2.today_dir = lambda today: step1.today_dir(today, inbox_root)
        step2.insert_order = _fail
        try:
            result = step2.run(today="2026-08-14", send_email=False)
        finally:
            step2.today_dir = original_today_dir
            step2.insert_order = original_insert_order

        assert "실패 1" in result
        with open(os.path.join(output_dir, "2026-08-14_epost_result.csv"), encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        assert "실패" in rows[1][5]


def test_response_without_tracking_number_is_treated_as_failure_not_success():
    # 우체국 서버가 200 응답을 주면서도 regiNo가 빠진 비정상 케이스 — 성공(접수완료)으로
    # 잘못 기록되면 대표님이 실제로는 안 나간 물건을 나간 걸로 착각할 수 있음.
    order = build_standard_order(
        Channel.PHONE_TEXT, "phone-1", ProductGroup.YUJACHEONG,
        "이동훈", "01090291152", "서울 마포구 ...", "04021", 1,
    )

    with tempfile.TemporaryDirectory() as inbox_root:
        output_dir = _write_pending(inbox_root, "2026-08-14", [order])
        _set_epost_env()

        original_today_dir = step2.today_dir
        original_insert_order = step2.insert_order
        step2.today_dir = lambda today: step1.today_dir(today, inbox_root)
        step2.insert_order = lambda auth_key, security_key, params: {"resultCode": "OK"}  # regiNo 없음
        try:
            result = step2.run(today="2026-08-14", send_email=False)
        finally:
            step2.today_dir = original_today_dir
            step2.insert_order = original_insert_order

        assert "실패 1" in result
        with open(os.path.join(output_dir, "2026-08-14_epost_result.csv"), encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        assert rows[1][4] == ""  # 송장번호 없음
        assert "실패" in rows[1][5]
        assert "접수완료" not in rows[1][5]


def test_smartstore_orders_produce_naver_dispatch_excel():
    order = build_standard_order(
        Channel.SMARTSTORE, "2026082512345", ProductGroup.CHEONGYUJA,
        "정다은", "01099998888", "전남 고흥군 ...", "59554", 3,
    )

    with tempfile.TemporaryDirectory() as inbox_root:
        output_dir = _write_pending(inbox_root, "2026-08-14", [order])
        _set_epost_env()

        original_today_dir = step2.today_dir
        original_insert_order = step2.insert_order
        step2.today_dir = lambda today: step1.today_dir(today, inbox_root)
        step2.insert_order = lambda auth_key, security_key, params: {"regiNo": "6890166064699"}
        try:
            step2.run(today="2026-08-14", send_email=False)
        finally:
            step2.today_dir = original_today_dir
            step2.insert_order = original_insert_order

        dispatch_path = os.path.join(output_dir, "2026-08-14_네이버발송처리.xlsx")
        assert os.path.exists(dispatch_path)
        wb = openpyxl.load_workbook(dispatch_path)
        ws = wb["발송처리"]
        assert [c.value for c in ws[2]] == ["2026082512345", "택배,등기,소포", "우체국택배", "6890166064699"]


def test_second_run_same_day_accumulates_csv_and_dispatch_file():
    order1 = build_standard_order(
        Channel.SMARTSTORE, "id-1", ProductGroup.CHEONGYUJA,
        "조용만", "01011112222", "전남 고흥군 ...", "59554", 3,
    )
    order2 = build_standard_order(
        Channel.SMARTSTORE, "id-2", ProductGroup.CHEONGYUJA,
        "송호연", "01033334444", "전남 고흥군 ...", "59554", 3,
    )

    with tempfile.TemporaryDirectory() as inbox_root:
        _set_epost_env()
        original_today_dir = step2.today_dir
        original_insert_order = step2.insert_order
        step2.today_dir = lambda today: step1.today_dir(today, inbox_root)
        try:
            output_dir = _write_pending(inbox_root, "2026-08-14", [order1])
            step2.insert_order = lambda auth_key, security_key, params: {"regiNo": "111"}
            step2.run(today="2026-08-14", send_email=False)

            _write_pending(inbox_root, "2026-08-14", [order2])
            step2.insert_order = lambda auth_key, security_key, params: {"regiNo": "222"}
            step2.run(today="2026-08-14", send_email=False)
        finally:
            step2.today_dir = original_today_dir
            step2.insert_order = original_insert_order

        with open(os.path.join(output_dir, "2026-08-14_epost_result.csv"), encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        assert [r[0] for r in rows[1:]] == ["id-1", "id-2"]
        assert sum(1 for r in rows if r and r[0] == "원주문번호") == 1

        wb = openpyxl.load_workbook(os.path.join(output_dir, "2026-08-14_네이버발송처리.xlsx"))
        ws = wb["발송처리"]
        assert [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)] == ["id-1", "id-2"]


def test_naver_api_dispatch_success_skips_excel_generation():
    order = build_standard_order(
        Channel.SMARTSTORE, "2026082512345", ProductGroup.CHEONGYUJA,
        "정다은", "01099998888", "전남 고흥군 ...", "59554", 3,
    )

    with tempfile.TemporaryDirectory() as inbox_root:
        output_dir = _write_pending(inbox_root, "2026-08-14", [order])
        _set_epost_env()

        original_today_dir = step2.today_dir
        original_insert_order = step2.insert_order
        original_naver_client = step2.NaverCommerceClient
        original_dispatch = step2.dispatch_orders
        step2.today_dir = lambda today: step1.today_dir(today, inbox_root)
        step2.insert_order = lambda auth_key, security_key, params: {"regiNo": "6890166064699"}
        step2.NaverCommerceClient = lambda client_id, client_secret: object()
        step2.dispatch_orders = lambda client, mapping: {
            "success_ids": list(mapping.keys()), "fail_infos": [],
        }

        old_env = {k: os.environ.get(k) for k in ("NAVER_CLIENT_ID", "NAVER_CLIENT_SECRET")}
        os.environ["NAVER_CLIENT_ID"] = "id"
        os.environ["NAVER_CLIENT_SECRET"] = "secret"
        try:
            result = step2.run(today="2026-08-14", send_email=False)
        finally:
            step2.today_dir = original_today_dir
            step2.insert_order = original_insert_order
            step2.NaverCommerceClient = original_naver_client
            step2.dispatch_orders = original_dispatch
            for k, v in old_env.items():
                if v is None:
                    os.environ.pop(k, None)
                else:
                    os.environ[k] = v

        assert "네이버 발송처리 API로 1건 자동 등록 완료" in result
        # API로 이미 등록됐으니 업로드용 엑셀을 따로 첨부/안내하지 않아야 함
        assert "엑셀 일괄발송" not in result


def test_naver_api_dispatch_failure_falls_back_to_excel():
    order = build_standard_order(
        Channel.SMARTSTORE, "2026082512345", ProductGroup.CHEONGYUJA,
        "정다은", "01099998888", "전남 고흥군 ...", "59554", 3,
    )

    with tempfile.TemporaryDirectory() as inbox_root:
        output_dir = _write_pending(inbox_root, "2026-08-14", [order])
        _set_epost_env()

        original_today_dir = step2.today_dir
        original_insert_order = step2.insert_order
        original_naver_client = step2.NaverCommerceClient
        original_dispatch = step2.dispatch_orders
        step2.today_dir = lambda today: step1.today_dir(today, inbox_root)
        step2.insert_order = lambda auth_key, security_key, params: {"regiNo": "6890166064699"}
        step2.NaverCommerceClient = lambda client_id, client_secret: object()

        def _fail_dispatch(client, mapping):
            raise RuntimeError("네이버 서버 오류")
        step2.dispatch_orders = _fail_dispatch

        old_env = {k: os.environ.get(k) for k in ("NAVER_CLIENT_ID", "NAVER_CLIENT_SECRET")}
        os.environ["NAVER_CLIENT_ID"] = "id"
        os.environ["NAVER_CLIENT_SECRET"] = "secret"
        try:
            result = step2.run(today="2026-08-14", send_email=False)
        finally:
            step2.today_dir = original_today_dir
            step2.insert_order = original_insert_order
            step2.NaverCommerceClient = original_naver_client
            step2.dispatch_orders = original_dispatch
            for k, v in old_env.items():
                if v is None:
                    os.environ.pop(k, None)
                else:
                    os.environ[k] = v

        assert "엑셀로 대체" in result
        assert os.path.exists(os.path.join(output_dir, "2026-08-14_네이버발송처리.xlsx"))


if __name__ == "__main__":
    tests = [obj for name, obj in list(globals().items()) if name.startswith("test_")]
    for t in tests:
        t()
        print(f"PASS {t.__name__}")
    print(f"\n{len(tests)} tests passed")
