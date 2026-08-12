import os
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import openpyxl  # noqa: E402

from courier_tier import Package  # noqa: E402
from naver_dispatch_export import write_naver_dispatch_excel  # noqa: E402
from schema import Channel, ProductGroup, build_standard_order  # noqa: E402


def _order(channel, order_id, name="조용만"):
    return build_standard_order(
        channel, order_id, ProductGroup.YUJACHEONG, name, "01011112222", "주소", "11111", 2,
        product_detail="로우슈거",
    )


def test_writes_smartstore_rows_with_real_tracking_numbers():
    results = [
        (Package(_order(Channel.SMARTSTORE, "2026080867140091", "조용만"), "유자청 로우슈거", "2kg"),
         {"regiNo": "6890161633630"}),
        (Package(_order(Channel.SMARTSTORE, "2026080760109111", "송호연"), "유자청 로우슈거", "2kg"),
         {"regiNo": "6890161633631"}),
    ]

    with tempfile.TemporaryDirectory() as tmp_dir:
        path = os.path.join(tmp_dir, "dispatch.xlsx")
        row_count = write_naver_dispatch_excel(results, path)

        assert row_count == 2
        wb = openpyxl.load_workbook(path)
        ws = wb["발송처리"]
        assert [c.value for c in ws[1]] == ["상품주문번호", "배송방법", "택배사", "송장번호"]
        assert [c.value for c in ws[2]] == ["2026080867140091", "택배,등기,소포", "우체국택배", "6890161633630"]
        assert [c.value for c in ws[3]] == ["2026080760109111", "택배,등기,소포", "우체국택배", "6890161633631"]


def test_non_smartstore_orders_are_excluded():
    results = [
        (Package(_order(Channel.WHOLESALE, "id-1"), "유자청", "2kg"), {"regiNo": "6890161600000"}),
    ]
    with tempfile.TemporaryDirectory() as tmp_dir:
        path = os.path.join(tmp_dir, "dispatch.xlsx")
        row_count = write_naver_dispatch_excel(results, path)
        assert row_count == 0


def test_test_mode_tracking_numbers_are_excluded():
    results = [
        (Package(_order(Channel.SMARTSTORE, "id-1"), "유자청", "2kg"), {"regiNo": "TESTREGINOAPI"}),
    ]
    with tempfile.TemporaryDirectory() as tmp_dir:
        path = os.path.join(tmp_dir, "dispatch.xlsx")
        row_count = write_naver_dispatch_excel(results, path)
        assert row_count == 0


if __name__ == "__main__":
    tests = [obj for name, obj in list(globals().items()) if name.startswith("test_")]
    for t in tests:
        t()
        print(f"PASS {t.__name__}")
    print(f"\n{len(tests)} tests passed")
